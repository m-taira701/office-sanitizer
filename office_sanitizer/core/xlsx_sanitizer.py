# office_sanitizer/core/xlsx_sanitizer.py
from __future__ import annotations

import logging
import os
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection

from .base import Sanitizer, CommonSanitizeOptions
from .utils import (
    iter_office_files,
    resolve_output_path,
    process_with_temp_copy,
    zip_sanitize_docprops,
)

logger = logging.getLogger(__name__)

@dataclass(frozen=True)
class ExcelSanitizeOptions(CommonSanitizeOptions):
    """
    Excelサニタイズオプション
    
    Attributes:
        zoom (int): 表示倍率 (デフォルト: 100)
        focus_cell (str): 選択セル (デフォルト: A1)
        set_first_sheet_active (bool): 最初のシートをアクティブにするかどうか。
        remove_comments (bool): コメントを削除するかどうか。
    """
    zoom: int = 100
    focus_cell: str = "A1"
    set_first_sheet_active: bool = True
    remove_comments: bool = True


class XlsxSanitizer(Sanitizer):
    def sanitize(self, path: str | os.PathLike, options: ExcelSanitizeOptions = ExcelSanitizeOptions()) -> None:
        """
        Excelファイル(.xlsx)をサニタイズします。
        
        Args:
            path: ファイルまたはディレクトリのパス
            options: サニタイズオプション (ExcelSanitizeOptions)
        """
        p = Path(path)

        targets: list[Path]
        if p.is_dir():
            targets = list(iter_office_files(p, "*.xlsx", recursive=options.recursive, skip_prefix="~$"))
        elif p.is_file():
            targets = [p]
        else:
            raise FileNotFoundError(f"Path not found: {p}")

        for f in targets:
            try:
                self._sanitize_one_xlsx(f, options)
                logger.info(f"サニタイズ完了: {f.name}")
            except Exception as e:
                logger.error(f"サニタイズ失敗: {f.name} - {str(e)}")

    def _sanitize_one_xlsx(self, src: Path, options: ExcelSanitizeOptions) -> None:
        if src.suffix.lower() != ".xlsx":
            return

        dst = resolve_output_path(src, options.in_place, options.output_dir, ".sanitized.xlsx")

        def processor(tmp_workbook_path: Path) -> None:
            # 1) openpyxl pass: zoom, selection, comments, basic properties
            _openpyxl_sanitize(tmp_workbook_path, options)

            # 2) zip-level pass: aggressively blank docProps/* metadata
            if options.remove_metadata:
                zip_sanitize_docprops(tmp_workbook_path)

        process_with_temp_copy(src, dst, processor)


def _openpyxl_sanitize(xlsx_path: Path, options: ExcelSanitizeOptions) -> None:
    # data_only=False to preserve formulas etc.
    wb = load_workbook(filename=str(xlsx_path), data_only=False, keep_vba=False)
    tmp_path = xlsx_path.with_suffix(".openpyxl.xlsx")
    saved = False
    try:
        # Active sheet -> first sheet
        if options.set_first_sheet_active and wb.sheetnames:
            wb.active = 0

        # Remove comments (cell comments + (best-effort) threaded comments)
        if options.remove_comments:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.comment is not None:
                            cell.comment = None

                if hasattr(ws, "_comments"):
                    try:
                        ws._comments = []  # type: ignore[attr-defined]
                    except Exception:
                        pass

        # Zoom + focus cell
        focus = options.focus_cell
        for ws in wb.worksheets:
            sv = ws.sheet_view
            try:
                sv.zoomScale = int(options.zoom)
            except Exception:
                pass

            try:
                sv.topLeftCell = focus
            except Exception:
                pass

            try:
                if not sv.selection:
                    sv.selection = [Selection(activeCell=focus, sqref=focus)]
                else:
                    sv.selection[0].activeCell = focus
                    sv.selection[0].sqref = focus
            except Exception:
                pass

        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass
        wb.save(filename=str(tmp_path))
        saved = True
        
    finally:
        wb.close()
        # Handle temp file cleanup and replacement
        if saved:
            try:
                tmp_path.replace(xlsx_path)
            except Exception:
                if tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except Exception:
                        pass
        elif tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass
