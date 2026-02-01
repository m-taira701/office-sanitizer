# office_sanitizer/excel.py
from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection

from .common import CommonSanitizeOptions, iter_office_files, process_with_temp_copy, resolve_output_path, zip_sanitize_docprops

@dataclass(frozen=True)
class ExcelSanitizeOptions(CommonSanitizeOptions):
    # TODO 外部変数化
    zoom: int = 100
    focus_cell: str = "A1"
    set_first_sheet_active: bool = True
    remove_comments: bool = True  # cell comments / threaded comments


def sanitize_excel(path: str | os.PathLike, options: ExcelSanitizeOptions = ExcelSanitizeOptions()) -> None:
    """
    Sanitize .xlsx files:
      - Zoom -> options.zoom
      - Active cell -> options.focus_cell (A1)
      - Active sheet -> first sheet (optional)
      - Remove comments (optional)
      - Remove metadata (docProps/core.xml, docProps/custom.xml, docProps/app.xml) (optional)

    Accepts a file or directory. Directory mode finds *.xlsx (skips temp "~$" files).
    """
    p = Path(path)

    targets: list[Path]
    if p.is_dir():
        targets = list(iter_office_files(p, "*.xlsx", recursive=options.recursive, skip_prefix="~$"))
    elif p.is_file():
        targets = [p]
    else:
        raise FileNotFoundError(f"Path not found: {p}")

    for f in targets:
        _sanitize_one_xlsx(f, options)


def _sanitize_one_xlsx(src: Path, options: ExcelSanitizeOptions) -> None:
    if src.suffix.lower() != ".xlsx":
        return

    dst = resolve_output_path(src, options.in_place, options.output_dir, ".sanitized.xlsx")

    def processor(tmp_workbook_path: Path) -> None:
        # 1) openpyxl pass: zoom, selection, comments, basic properties
        _openpyxl_sanitize(tmp_workbook_path, options)

        # 2) zip-level pass: aggressively blank docProps/* metadata
        if options.remove_metadata:
            zip_sanitize_docprops(tmp_workbook_path)

    process_with_temp_copy(src, dst, processor)


def _openpyxl_sanitize(xlsx_path: Path, options: ExcelSanitizeOptions) -> None:
    # data_only=False to preserve formulas etc.
    wb = load_workbook(filename=str(xlsx_path), data_only=False, keep_vba=False)
    tmp_path = xlsx_path.with_suffix(".openpyxl.xlsx")
    saved = False
    try:
        # Active sheet -> first sheet
        if options.set_first_sheet_active and wb.sheetnames:
            wb.active = 0

        # Remove comments (cell comments + (best-effort) threaded comments)
        if options.remove_comments:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.comment is not None:
                            cell.comment = None

                if hasattr(ws, "_comments"):
                    try:
                        ws._comments = []  # type: ignore[attr-defined]
                    except Exception:
                        pass

        # Zoom + focus cell
        focus = options.focus_cell
        for ws in wb.worksheets:
            sv = ws.sheet_view
            try:
                sv.zoomScale = int(options.zoom)
            except Exception:
                pass

            try:
                sv.topLeftCell = focus
            except Exception:
                pass

            try:
                if not sv.selection:
                    sv.selection = [Selection(activeCell=focus, sqref=focus)]
                else:
                    sv.selection[0].activeCell = focus
                    sv.selection[0].sqref = focus
            except Exception:
                pass

        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass
        wb.save(filename=str(tmp_path))
        saved = True
        
    finally:
        wb.close()
        if saved:
            try:
                tmp_path.replace(xlsx_path)
            except Exception:
                if tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except Exception:
                        pass
        elif tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass


